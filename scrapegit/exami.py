from selenium import webdriver
import docx
import openpyxl
#start PhantomJS
service_args = [ '--proxy=localhost:9150', '--proxy-type=socks5', ]
driver = webdriver.PhantomJS(service_args=service_args)
#doc = docx.Document()
lis = ['http://www.studyguideindia.com/Exams/engineering-entrance-exams-all-india-level.asp', 
	   'http://www.studyguideindia.com/Exams/medical-and-releated-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/dental-entrance-exams-state-level.asp',
	   'http://www.studyguideindia.com/Exams/pharmacy-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/fishery-sciences-entrance-exams.asp',
	   'http://www.studyguideindia.com/Exams/space-marine-architecture-aviation-and-information-technology-exams.asp',
	   'http://www.studyguideindia.com/Exams/civil-services-entrance-exam.asp', 
	   'http://www.studyguideindia.com/Exams/fashion-technology-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/alternate-medicine-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/law-entrance-examination.asp', 
	   'http://www.studyguideindia.com/Exams/physiotherapy-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/nursing-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/mbbs-bds-related-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/engineering-entrance-exams-state-level.asp', 
	   'http://www.studyguideindia.com/Exams/veterinary-sciences-entrance-exams.asp',
	   'http://www.studyguideindia.com/Exams/bio-technology-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/defence-education-entrance-exam.asp', 
	   'http://www.studyguideindia.com/Exams/mass-and-media-entrance-exams.asp', 
	   'http://www.studyguideindia.com/Exams/hospitality-travel-tourism-exams.asp',
	   'http://www.studyguideindia.com/Exams/management-entrance-examination.asp',
	   'http://www.studyguideindia.com/Exams/Competitive-Exams.asp']
xl = openpyxl.Workbook()
sh = xl.active
i = 1
for linkss in lis:
	driver.get(linkss)
	print '(+) Scraping Link : ' + linkss
	print '(+) Rows Count : ' + str(i)
	try:
		div = driver.find_element_by_class_name('listdt') 
	except:
		div = driver.find_element_by_class_name('PB6') 
	li = div.find_elements_by_tag_name('li')
	print '(+) Exams Found : ' + str(len(li))
	for aa in li:
		sh.cell(row=i, column=1).value = i
		try:
			a = aa.find_element_by_tag_name('a')
			sh.cell(row=i, column=2).value = a.get_attribute('href')
			sh.cell(row=i, column=3).value = a.text
		except:
			sh.cell(row=i, column=2).value = aa.text
		i += 1
		xl.save('exmli.xlsx')
driver.close		