# from instabot import Bot
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import time
import os
import openpyxl
from random import randint, choice
import fix


def file_rename(newname, folder_of_download, time_to_wait=120):
    time_counter = 0
    try :
        filename = max([f for f in os.listdir(folder_of_download)], key=lambda xa :   os.path.getctime(os.path.join(folder_of_download,xa)))
        while '.crdownload' in filename:
            # print(filename)
            time.sleep(1)
            filename = max([f for f in os.listdir(folder_of_download)], key=lambda xa :   os.path.getctime(os.path.join(folder_of_download,xa)))
            time_counter += 1
            if time_counter > time_to_wait:
                return filename
        filename = max([f for f in os.listdir(folder_of_download)], key=lambda xa :   os.path.getctime(os.path.join(folder_of_download,xa)))
    except :
        filename = ''
    # os.rename(os.path.join(folder_of_download, filename), os.path.join(folder_of_download, newname))
    return filename


def file_move() :
    xl = openpyxl.load_workbook('/home/diablo/pros/personal/px/media/final_scac1.xlsx')
    al = xl.active
    ff = os.listdir('/home/diablo/pros/personal/px/media/nmedia/')
    floc = []
    for f in ff :
        mf = False
        floc.append({f:'','sub':''})
        for i in range(1,al.max_row+1) :
            n = al.cell(row=i,column=5).value
            cl = 6
            if int(n) > 0 :
                cl += int(n)
            for v in range(6,cl+1) :
                if al.cell(row=i,column=v).value == f :
                    mf = True
                    s = al.cell(row=i,column=1).value
                    s = ''.join(e for e in s if e.isalnum())
                    floc[-1][f] = s
                    if int(n) > 1 :
                        floc[-1]['sub'] = al.cell(row=i,column=4).value
                    break
            if mf :
                break
        if mf == False:
            continue
        if os.path.isdir('/home/diablo/pros/personal/px/media/nmediafn/'+floc[-1][f]) == False :
            os.makedirs('/home/diablo/pros/personal/px/media/nmediafn/'+floc[-1][f])
        pt = '/home/diablo/pros/personal/px/media/nmediafn/'+floc[-1][f]+'/'
        if floc[-1]['sub'] != '' :
            if os.path.isdir('/home/diablo/pros/personal/px/media/nmediafn/'+floc[-1][f]+'/'+floc[-1]['sub']) == False :
                os.makedirs('/home/diablo/pros/personal/px/media/nmediafn/'+floc[-1][f]+'/'+floc[-1]['sub'])
            pt += floc[-1]['sub']+'/'
        os.rename("/home/diablo/pros/personal/px/media/nmedia/"+f, pt+f)
        # os.replace("/home/diablo/pros/personal/px/media/mm2/"+f, pt+f)
        # shutil.move("/home/diablo/pros/personal/px/media/mm2/"+f, pt+f)
        # print(pt)

    print(floc)

def find_file_locs(name, path):
    result = []
    for root, dirs, files in os.walk(path):
        if name in files:
            result.append(os.path.join(root, name))
    return result


def get_file(acc) :
    path = fix.PATH
    path += fix.FPATHS[acc]
    fs = os.listdir(path)
    if '.xlsx' in fs[0] :
        return path+fs[1]
    else :
        return path+fs[0]

def get_official_cap(fl) :
    fl2 = fl
    fl = fl.split('/')[-1]
    path = fix.PATH + 'off/final_scac1.xlsx'
    path2 = fix.PATH + 'off/uploaded/uploads.xlsx'
    uxl = openpyxl.load_workbook(path2)
    uxla = uxl.active
    capxl = openpyxl.load_workbook(path)
    capact = capxl.active
    rc = uxla.max_row+1
    if 'cj' in fl2 : xx = 'cj'
    if 'escon' in fl2 : xx = 'escon'
    if 'plus2' in fl2 : xx = 'plus2'
    for i in range(1,capact.max_row+1) :
        n = capact.cell(row=i,column=5).value
        cl = 6
        if int(n) > 0 :
            cl += int(n)
        for v in range(6,cl+1) :
            if capact.cell(row=i,column=v).value == fl :
                uxla.cell(row=rc,column=1).value = capact.cell(row=i,column=1).value
                uxla.cell(row=rc,column=2).value = capact.cell(row=i,column=2).value
                uxla.cell(row=rc,column=3).value = capact.cell(row=i,column=3).value
                uxla.cell(row=rc,column=4).value = capact.cell(row=i,column=4).value
                uxla.cell(row=rc,column=4).value = capact.cell(row=i,column=v).value
                uxl.save(fix.PATH + 'off/uploaded/uploads.xlsx')
                x = capact.cell(row=i,column=3).value
                try :
                    y = x.split(' ')
                except :
                    return 'Follow for more... \n.\n.\n.\n.\n.\n.\n.' + fix.OFFICIAL_TAGS[xx]
                z = [j for j in y if j.startswith('@') == False]
                z = ' '.join(z)
                # z = z.split(' ')
                # z = [j for j in z if j.isalnum() == True or '#' in j]
                # z = ' '.join(z)
                if '#' not in z :
                    z += '\n.\n.\n.\n.' + fix.OFFICIAL_TAGS[xx]         
                return z
    
    return 'Follow for more....\n.\n.\n.\n.\n.' + fix.OFFICIAL_TAGS[xx]

def get_caption(acc, tag=False) :
    path = fix.PATH
    path += fix.FPATHS[acc]
    path += 'caption.xlsx'

    ox = openpyxl.load_workbook(path)
    oy = ox.active
    if oy.cell(row=oy.max_row,column=1).value != None :
        cap = oy.cell(row=oy.max_row,column=1).value
        oy.cell(row=oy.max_row,column=1).value = None
        ox.save(path)
    else :
        cap = '❤️'
    if tag :
        path2 = fix.PATH+'tags.xlsx'
        ox = openpyxl.load_workbook(path2)
        oy = ox.active
        tgs = {}
        for t in range(2,oy.max_row+1) :
            tgs[oy.cell(row=t,column=1).value] = oy.cell(row=t,column=2).value
        tags = tgs[acc]
        cap += '\n.\n.\n.\n.\n.\n'
        cap += tags
    return cap

def del_file(fl, acc_typ='off') :
    # os.remove(fl)
    if acc_typ == 'off' :
        if 'cj' in fl :
            pt = '/home/diablo/pros/personal/px/media/off/uploaded/cj/'
        if 'escon' in fl :
            pt = '/home/diablo/pros/personal/px/media/off/uploaded/escon/'
        if 'plus2' in fl :
            pt = '/home/diablo/pros/personal/px/media/off/uploaded/plus2/'
    f = fl.split('/')[-1]
    os.rename(fl, pt+f)
    print('File Moved : '+fl)

def get_comment(acc,typ) :
    if typ == 'unof' :
        pass
    else :
        pass
    return choice(fix.COMCH)

def sel_login(unm,pw) :
    options = Options()
    # options.add_argument('--headless=new')
    dr = webdriver.Chrome(options=options)
    dr.get('https://instagram.com')
    dr.maximize_window()
    time.sleep(5)
    inp = dr.find_elements(By.CLASS_NAME, "_aa4b")
    inp[0].send_keys(unm)
    inp[1].send_keys(pw)
    dr.find_elements(By.CLASS_NAME,"_acap")[0].click()
    time.sleep(10)  
    try :
        dr.find_elements(By.CLASS_NAME, "_ap36")[-1].click()
        time.sleep(5)
    except :
        pass
    try :
        dr.find_elements(By.CLASS_NAME, "_a9_1")[0].click()
    except :
        pass
    print('Logged in username : '+unm)
    time.sleep(5)
    return dr

def scrape_data(acc, dr) :
    path = fix.PATH
    path += 'trial.xlsx'
    ox = openpyxl.load_workbook(path)
    oy = ox.active
    for x in acc :
        url = 'https://www.instagram.com/'+x+'/'
        dr.get(url)
        time.sleep(20)
        print('Account Loaded : ' + x)
        print('----')
        time.sleep(3)
        cls = '_aagu'
        posts = dr.find_elements(By.CLASS_NAME, cls)
        posts[1].click()
        t = 1
        while True :
            time.sleep(5)
            txt = dr.find_elements(By.CLASS_NAME, '_a9zs')
            if len(txt) >= 1 :
                cp = txt[0].text
            else :
                cp = ''
            i = oy.max_row
            oy.cell(row=i+1,column=1).value = x
            oy.cell(row=i+1,column=2).value = dr.current_url
            oy.cell(row=i+1,column=3).value = cp
            ox.save(path)
            print('Post Number : ' + str(t))
            print('----------------')
            t += 1
            ncl = dr.find_elements(By.CLASS_NAME, '_abl-')
            try :
                ncl[1].click()
            except :
                break

        # while acp < plen :
        #     for t in range(acp, plen) :
        #         posts[t].click()
        #         time.sleep(5)
        #         txt = dr.find_elements(By.CLASS_NAME, '_a9zs')
        #         if len(txt) >= 1 :
        #             cp = txt[0].text
        #         else :
        #             cp = ''
        #         i = oy.max_row
        #         oy.cell(row=i+1,column=1).value = x
        #         oy.cell(row=i+1,column=2).value = dr.current_url
        #         oy.cell(row=i+1,column=3).value = cp
        #         ox.save(path)
        #         print('Post Number : ' + str(t))
        #         print('----------------')
        #         dr.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
        #         acp += 1
            
        #     posts = dr.find_elements(By.CLASS_NAME, cls)
        #     plen = len(posts)

        #     # if plen == (acp+1) :
        #     time.sleep(30)

        #     posts = dr.find_elements(By.CLASS_NAME, cls)
        #     plen = len(posts)
        #     if acp == plen :
        #         time.sleep(30)
            # posts[plen-1].click()


        # while True :
        #     posts = dr.find_elements(By.CLASS_NAME, cls)
        #     plen = len(posts)
        #     for t in range(plen) :
        #         if acp 
        #         posts[acp].click()
        #         acp += 1

        #     if acp == (plen-1) :
        #         time.sleep(30)
        #         break

        # for t in range(200) :
        #     time.sleep(3)
        #     cls = '_aagu'
        #     posts = dr.find_elements(By.CLASS_NAME, cls)
        #     plen = len(posts)
        #     acp = 0
        #     if t == (plen-1) :
        #         pass
        #     try :
        #         posts[t].click()
        #     except :
        #         time.sleep(20)
        #         posts = dr.find_elements(By.CLASS_NAME, cls)
        #         try :
        #             posts[t-1].click()
        #             time.sleep(10)
        #             posts[t].click()
        #         except :
        #             time.sleep(20)
        #             posts = dr.find_elements(By.CLASS_NAME, cls)
        #             try :
        #                 posts[t-1].click()
        #                 time.sleep(10)
        #                 posts[t].click()
        #             except :
        #                 break
        #     time.sleep(5)
        #     txt = dr.find_elements(By.CLASS_NAME, '_a9zs')
        #     if len(txt) >= 1 :
        #         cp = txt[0].text
        #     else :
        #         cp = ''
        #     i = oy.max_row
        #     oy.cell(row=i+1,column=1).value = x
        #     oy.cell(row=i+1,column=2).value = dr.current_url
        #     oy.cell(row=i+1,column=3).value = cp
        #     ox.save(path)
        #     print('Post Number : ' + str(t))
        #     print('----------------')
        #     dr.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)

def like_list_follow(dr) :
    lk = dr.find_elements(By.CLASS_NAME, 'x18hxmgj')
    p = None
    for v in lk :
        if 'likes' in v.text : p = v
    if p is not None : p.click()
    time.sleep(5)
    i = -1
    while True :
        y = dr.find_elements(By.CLASS_NAME, '_aaco')
        try :
            if y[i].text == 'Follow' : y[i].click()
        except :
            break
        time.sleep(randint(3,8))
        i = i-1
        if i < -6 :
            break
    print('-----New Thread------')

def follower(tag,typ,dr) :
    print('Type : '+typ)
    print('Data : '+tag)
    if typ == 'tag' :
        url = 'https://www.instagram.com/explore/tags/'+tag+'/'
    elif typ == 'account' :
        url = 'https://www.instagram.com/'+tag+'/'
    elif typ == 'post' :
        url = 'https://www.instagram.com/p/'+tag+'/'
    for t in range(2,6) :
        dr.get(url)
        time.sleep(20)
        if typ == 'tag' or typ == 'account' :
            cls = '_aagu'
        if typ != 'post' :
            posts = dr.find_elements(By.CLASS_NAME, cls)
            try :
                posts[t].click()
            except :
                continue
            time.sleep(3)
        like_list_follow(dr)
    return dr

def unfollower(dr, unm) :
    while True :
        dr.get('https://www.instagram.com/' + unm + '/following/')
        time.sleep(10)
        y = dr.find_elements(By.CLASS_NAME, '_aaco')
        for i in range(len(y)) :
            if y[i].text == 'Following' : 
                y[i].click()
                time.sleep(2)
                dr.find_elements(By.CLASS_NAME, '_a9--')[0].click()
                time.sleep(3)
        flag = True
        for x in y :
            if x.text == 'Following' : flag = False
        if flag : break
    return dr

def comment_like(typ, tag, comment, dr, cnum) :
    if typ == 'tag' :
        url = 'https://www.instagram.com/explore/tags/'+tag+'/'
    elif typ == 'account' :
        url = 'https://www.instagram.com/'+tag+'/'
    dr.get(url)
    time.sleep(10)
    for t in range(20) :
        posts = dr.find_elements(By.CLASS_NAME, '_aagu')
        try :
            posts[t].click()
        except :
            dr.refresh()
            time.sleep(10)
            posts = dr.find_elements(By.CLASS_NAME, '_aagu')
            try :
                posts[t].click()
            except :
                break
        time.sleep(5)
        try :
            pos = dr.find_element(By.CLASS_NAME, '_aamw')
        except :
            continue
        pos.click()
        time.sleep(3)
        if comment != '' and cnum%3 == 0 :
            ti = dr.find_elements(By.CLASS_NAME, '_akhn')
            if len(ti) == 0 :
                continue
            ti[0].find_element(By.CLASS_NAME,'x1lliihq').click()
            time.sleep(3)
            ec = dr.find_elements(By.CLASS_NAME,'_ajr2')
            for _ in range(choice(range(1,7))) :
                time.sleep(0.5)
                try :
                    ec[choice(range(1,15))].click()
                except :
                    continue
            time.sleep(2)
            tip = dr.find_element(By.CLASS_NAME, '_am-5')
            tip.click()
            time.sleep(10)
            print('Commented')
        cnum += 1
        es = dr.find_element(By.TAG_NAME, 'body')
        es.send_keys(Keys.ESCAPE)
        print('Tag : '+ tag +'  |  Liked Post : '+str(t))
        time.sleep(2)
    return (dr,cnum)

def auto_upload(file, dr, caption) :
    print('-------------Upload Started----------------')
    time.sleep(5)
    inp1 = dr.find_elements(By.CLASS_NAME, "x5n08af")
    for x in inp1 :
        if x.text == 'Create' :
            x.click()
            break
    time.sleep(5)
    inp1 = dr.find_elements(By.CLASS_NAME, "x9f619")
    for x in inp1 :
        if x.text == 'Post' :
            x.click()
            break
    time.sleep(5)
    cr_clk = dr.find_element(By.CLASS_NAME, "_ac69")
    cr_clk.send_keys(file)
    time.sleep(20)
    try :
        ok_clk2 = dr.find_elements(By.CLASS_NAME, "xdko459")
        for x in ok_clk2 :
            if x.text == 'OK' :
                x.click()
                time.sleep(2)
                break
    except :
        pass
    k = dr.find_elements(By.CLASS_NAME, "_ap30")
    k[1].click()
    time.sleep(2)
    m = dr.find_elements(By.CLASS_NAME, "x1nhvcw1")
    for x in m :
        if x.text == 'Original' :
                x.click()
                break
    time.sleep(2)
    ok_clk = dr.find_elements(By.CLASS_NAME, "_acan")
    for x in ok_clk :
        if x.text == 'OK' :
            x.click()
            break
    nxt_clk = dr.find_elements(By.CLASS_NAME, "x1i10hfl")
    for x in nxt_clk :
        if x.text == 'Next' :
            x.click()
            break
    time.sleep(5)
    nxt_clk = dr.find_elements(By.CLASS_NAME, "x1i10hfl")
    for x in nxt_clk :
        if x.text == 'Next' :
            x.click()
            break
    time.sleep(5)
    j = dr.find_elements(By.CLASS_NAME, "notranslate")
    cap = j[-1]
    try :
        cap.send_keys(caption)
    except :
        z = caption.split(' ')
        z = [j for j in z if j.isalnum() or '#' in j]
        try :
            cap.send_keys(' '.join(z))
        except :
            z = caption.split(' ')
            z = [j for j in z if j.isalnum()]
            cap.send_keys(' '.join(z))
    time.sleep(5)
    shr_clk = dr.find_elements(By.CLASS_NAME, "x1i10hfl")
    for x in shr_clk :
        if x.text == 'Share' :
            x.click()
            break
    time.sleep(30)
    if '.mp4' in file :
        time.sleep(120)
    print('-------------Upload Complete----------------')
    dr.find_element(By.TAG_NAME,'body').send_keys(Keys.ESCAPE)
    return dr

def one_cycle(driver=None, unm='', upload=False, fllw=False, unfllw=False, likr=False, acctp='unof', ftgs=None, ctgs=None) :
    # Upload Image
    if upload :
        f = get_file(unm)
        # cap = get_caption(unm,True)
        if acctp == 'off' :
            cap = get_official_cap(f)
        print('File Selected : '+f+'\nCaption : '+cap)
        driver = auto_upload(f, driver, cap)
        del_file(f)

    # Follow people
    if fllw :
        print('\n.\n.-------------------Starting Follower---------------------------')
        for x in ftgs :
            driver = follower(x,'tag', driver)
        print('\n.\n.-------------------Terminating Follower---------------------------')

    # Unfollow people
    if unfllw :
        print('\n.\n.-------------------Starting Unfollower---------------------------')
        driver = unfollower(driver,unm)
        print('\n.\n.-------------------Terminating Unfollower---------------------------')
    
    # Like and Comment on post
    if likr :
        print('\n.\n.-------------------Starting Liker / Commentor---------------------------')
        cnum = 1
        for tg in ctgs :
            cmnt = get_comment(unm,acctp)
            driver,cnum = comment_like('tag',tg,cmnt,driver,cnum)
        print('\n.\n.-------------------Terminating Liker / Commentor---------------------------')
    
    driver.close()
    time.sleep(60)

# UNOFFICIALS
for a in fix.ACCS :
    dr = sel_login(a['unm'],a['pw'])
    one_cycle(
        driver=dr,
        unm=a['unm'],
        upload=False,
        unfllw=False,
        fllw=True,
        likr=False,
        acctp='unof',
        ftgs=fix.FTAGS[a['unm']],
        ctgs=fix.CTAGS[a['unm']]
        )

# # OFFICIALS
# for a in fix.CRON_ACCS :
#     dr = sel_login(a['unm'],a['pw'])
#     one_cycle(
#         driver=dr,
#         unm=a['unm'],
#         upload=True,
#         unfllw=False,
#         fllw=False,
#         likr=False,
#         acctp='off',
#         ftgs=fix.FTAGS[a['unm']],
#         ctgs=fix.CTAGS[a['unm']]
#         )


########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################### SINGLE RUN SCRIPTS #########################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################
########################################################################################################################################################################

# dr = sel_login('plustwo.in','')
# # scrape_data(['learnwithbhu','business_idea_1513','discovering.ai','amigoscode','universeofprogrammers','dan_nanni','ai_junkies','cyber._.ox','o9go_','gamerebels','mindset.focused','chemistry_knowledgee','material_science007','success._.learning'],dr)
# scrape_data(['brgaming',],dr)

# x = openpyxl.load_workbook('/home/diablo/pros/personal/px/media/fscac2.xlsx')
# y = x.active
# # z = openpyxl.Workbook()
# # zx = z.active
# for t in range(1,y.max_row+1) :
#     p = y.cell(row=t,column=2).value
#     q = p.split('/')[-2]
#     y.cell(row=t,column=4).value = q
#     x.save('/home/diablo/pros/personal/px/media/fscac3.xlsx')

# x = openpyxl.load_workbook('/home/diablo/pros/personal/px/media/fscac3.xlsx')
# y = x.active
# dr = webdriver.Chrome()

# for t in range(1,y.max_row+1) :
#     dr.get('https://snapinsta.app/')
#     print('Row : '+str(t))
#     time.sleep(3)
#     url = y.cell(row=t,column=2).value
#     # if 'img_index' in url :
#     #     continue
#     inp = dr.find_element(By.ID, "url")
#     inp.send_keys(url)
#     time.sleep(1)
#     inp.send_keys(Keys.ENTER)
#     time.sleep(5)
#     a = dr.find_elements(By.CLASS_NAME, "download-media")
#     i = 0
#     k = 6
#     y.cell(row=t,column=5).value = len(a)
#     for aa in a :
#         nm = y.cell(row=t,column=4).value
#         if i > 0 :
#             nm += '_'+str(i)
#         time.sleep(2)
#         try :
#             aa.click()
#         except :
#             continue
#         try :
#             dr.find_element(By.CLASS_NAME, "close-button").click()
#         except :
#             pass
#         f = file_rename(nm,'/home/diablo/Downloads/')
#         y.cell(row=t,column=k).value = f
#         x.save('/home/diablo/pros/personal/px/media/final_scac1.xlsx')
#         k += 1
# print('____________________________________________________________')
# print('Completed download - Waiting to finish (10 minutes)')
# time.sleep(600)

# file_move()
    
# x = os.listdir('/home/diablo/pros/personal/px/media/off/cj')
# x += os.listdir('/home/diablo/pros/personal/px/media/off/escon')
# x += os.listdir('/home/diablo/pros/personal/px/media/off/plus2')
# x += os.listdir('/home/diablo/pros/personal/px/media/off/uploaded/cj')
# x += os.listdir('/home/diablo/pros/personal/px/media/off/uploaded/escon')
# x += os.listdir('/home/diablo/pros/personal/px/media/off/uploaded/plus2')

# z = []
# for y in x :
#     z.append(find_file_locs(y,'/home/diablo/pros/personal/px/media/bups'))

# for t in z :
#     if len(t) == 0 :
#         continue
#     for m in t :
#         os.remove(m)
#         print('removed : ' + m)

# print(z)

####################### CRON JOB ##########################

# def upload() :
#     # CJ
#     upload_img("plustwo.in" ,"", 
#                os.path.abspath(os.path.join(os.getcwd(), os.pardir)) + '/botscript/plustwo/caption.xlsx', 
#                os.path.abspath(os.path.join(os.getcwd(), os.pardir)) + '/botscript/plustwo/')
    # +2
    # upload_img("cyberjunkyard.in" ,"", 
    #            os.path.abspath(os.path.join(os.getcwd(), os.pardir)) + '/botscript/cj/caption.xlsx', 
    #            os.path.abspath(os.path.join(os.getcwd(), os.pardir)) + '/botscript/cj/')

# upload()

# def upload_img(username,password,xl,path) :
#     # Initialize bot and login
#     bot = Bot()
#     bot.login(username=username, password=password)

#     # Getting captions
#     ox = openpyxl.load_workbook(xl)
#     oy = ox.active
#     cap = {}
#     for t in range(2,oy.max_row+1) :
#         cap[oy.cell(row=t,column=1).value] = oy.cell(row=t,column=2).value

#     # Get image and caption
#     dr = os.listdir(path)
#     for x in dr :
#         img = path+x
#         if '.xlsx' in x :
#             continue
#         try :
#             caption = cap[x]
#             for t in range(2,oy.max_row+1) :
#                 if oy.cell(row=t,column=1).value == x :
#                     oy.delete_rows(t)
#                     break
#             ox.save(xl)
#         except :
#             if 'plustwo' in username :
#                 caption = 'Take the free personality test and get a free counselling session from the expert of your field. Click the link in bio or DM us. \n.\n.\n#career #counselling #counsellor #jee #neet #commerce #cbse #mentor #inspiration #motivation'
#             else :
#                 caption = 'Click the link in bio to know more. \n.\n.\n#cybersecurity #cyber #hacking #ethicalhacking #cybercrime'
        
#         # Upload media
#         bot.upload_photo(img, caption=caption)
#         os.remove(img)
#         break

# def follower() :
    # Search button
    # dr.find_elements(By.CLASS_NAME,"xr9ek0c")[2].click()
    # time.sleep(2)
    # # Search input and send tags
    # dr.find_elements(By.CLASS_NAME, "x1lugfcp")[0].send_keys(tag)
    # time.sleep(3)
    # # click on 1st result
    # dr.find_elements(By.CLASS_NAME, 'x1wdrske')[1].click()
    # time.sleep(5)