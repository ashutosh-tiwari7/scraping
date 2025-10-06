import os
import selenium
from selenium import webdriver
import openpyxl
import time
import random

pz = 0
prox = []
def rbro():
	br = webdriver.Chrome(r'/home/diablo/chromedriver')
	br.maximize_window()
	return br
def rebro():
	print('$_$ Setting Proxy $_$')
	pth = r'/home/diablo/chromedriver'
	global pz
	global prox
	if len(prox) == 0:
		dri = webdriver.Chrome(pth)
		dri.get('https://free-proxy-list.net/')
		proxt = dri.find_elements_by_xpath('//*[@id="proxylisttable"]/tbody/tr')
		for x in proxt:
			row_data = x.find_elements_by_tag_name('td')
			proxy = row_data[0].text+":"+row_data[1].text
			prox.append(proxy)
		dri.quit()	
	px = prox[pz]
	options = webdriver.ChromeOptions()
	options.add_argument("--start-maximized")
	options.add_argument('--proxy-server=http://%s' % px)

	driver = webdriver.Chrome(pth,options=options)
	pz += 1
	print('$_$ Proxy Changed : '+prox[pz])
	return driver


zl = os.listdir('/home/diablo/Desktop/scrpy/job/')
br = rbro()
for zlx in zl:
	zly = zlx.split('.')
	zxl = openpyxl.load_workbook('/home/diablo/Desktop/scrpy/job/'+zlx)
	zsh = zxl.active
	zc = zsh.max_row
	print('Workbook : ' + zlx + ' ; Rows : ' + str(zc))
	time.sleep(2)
	for zcx in range(1,zc+1):
		url = zsh.cell(row=zcx,column=2).value
		snm = zsh.cell(row=zcx,column=1).value
		br.get(url)
		print('Row : ' + str(zcx))
		print('URL : ' + url)
		i = 1
		xl1 = openpyxl.Workbook()
		sh1 = xl1.active

		a1  = br.find_element_by_tag_name('header')
		a2  = a1.find_element_by_tag_name('h1') #heading
		a3  = a1.find_element_by_tag_name('p') #Desc
		sh1.cell(row=i,column=1).value = a2.text
		sh1.cell(row=i,column=2).value = a3.text
		
		a4 = br.find_element_by_id('posts')
		a5 = a4.find_elements_by_class_name('post')

		a6 = a5[0] #work
		a6a = a6.find_elements_by_tag_name('li')
		if a6a:
			for a6b in a6a:
				sh1.cell(row=i,column=3).value = a6b.text
				i+=1
			i = 1
			a7 = a5[1] #activities
		else:
			a7 = a6
		
		a7a = a7.find_elements_by_tag_name('tr')
		for a7x in a7a:
			try:
				ax = a7x.find_elements_by_tag_name('td')
				ax1 = ax[0].find_element_by_tag_name('img')
				ax2 = ax[1].text
				ax3 = ax2.split('-')
				ax4 = '-'.join(ax3[1:])
				sh1.cell(row=i,column=4).value = ax3[0]
				sh1.cell(row=i,column=5).value = ax4
				sh1.cell(row=i,column=6).value = ax1.get_attribute('width')
				i+=1
			except:
				pass
		i = 1
		if not os.path.exists('/home/diablo/Desktop/scrpy/jobs/'+zly[0]):
			os.mkdir('/home/diablo/Desktop/scrpy/jobs/'+zly[0])
		try:
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm+'.xlsx')
		except:
			snm2 = snm.split('/')
			snm3 = '-'.join(snm2)
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm3+'.xlsx')
		time.sleep(random.randint(10,12))
		#br.get(url+'education')
		cla = br.find_element_by_id('fixed-head')
		cla.find_element_by_link_text('Education').click()
		try: 
			bxrx = br.find_element_by_id('hero9')
			if bxrx:
				br.quit()
				br = rbro()
				br.get(url+'education/')
		except:
			pass
		print('Section -- Education')
		b1 = br.find_element_by_id('posts')
		b2 = b1.find_elements_by_class_name('post')
		try:
			b3 = b2[1].find_element_by_tag_name('strong')
			sh1.cell(row=i,column=7).value = b3.text
		except:
			pass
		b2a = b2[0].find_elements_by_tag_name('li')
		for b2b in b2a:
			sh1.cell(row=i,column=8).value = b2b.text
			i+=1
		i = 1
		try:
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm+'.xlsx')
		except:
			snm2 = snm.split('/')
			snm3 = '-'.join(snm2)
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm3+'.xlsx')
		time.sleep(random.randint(10,12))
		#br.get(url+'skills')
		br.find_element_by_link_text('Skills').click()
		try:
			bxrx = br.find_element_by_id('hero9')
			if bxrx:
				br.quit()
				br = rbro()
				br.get(url+'skills')
		except:
			pass
		print('Section -- Skills')
		c1 = br.find_element_by_id('posts')
		c2 = br.find_elements_by_class_name('col-md-8')
		t = 9
		for c2x in c2:
			c3 = c2x.find_elements_by_tag_name('tr')
			for c4 in c3:
				try:
					cx = c4.find_elements_by_tag_name('td')
					cx1 = cx[0].find_element_by_tag_name('img')
					cx2 = cx[1].text
					cx3 = cx2.split('-')
					cx4 = '-'.join(cx3[1:])
					sh1.cell(row=i,column=t).value = cx3[0]
					sh1.cell(row=i,column=t+1).value = cx4
					sh1.cell(row=i,column=t+2).value = int(cx1.get_attribute('width'))*2
					i+=1
				except:
					pass
			i = 1
			t += 3
		try:
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm+'.xlsx')
		except:
			snm2 = snm.split('/')
			snm3 = '-'.join(snm2)
			xl1.save('/home/diablo/Desktop/scrpy/jobs/'+zly[0]+'/'+snm3+'.xlsx')
		print('Saved')
		time.sleep(random.randint(10,12))

