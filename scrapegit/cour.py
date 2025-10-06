import openpyxl
import selenium
from selenium import webdriver

br = webdriver.Chrome(r'/usr/local/bin/chromedriver')
wb = openpyxl.Workbook()
sh = wb.active
sli = ['fashion-and-interior-designing']

for xli in sli:
	br.get('https://targetstudy.com/courses/' + xli + '/')
	print '(+) Getting https://targetstudy.com/courses/' + xli + '/'
	tr = br.find_elements_by_tag_name('tr')
	print '(+)Rows found: ' + str(len(tr))
	i = 1

	for trx in tr:
		
		td = trx.find_elements_by_tag_name('td')
		if len(td) == 2:
			try:
				ar = td[0].find_element_by_tag_name('a')
				sh.cell(row=i, column=2).value = ar.text
				sh.cell(row=i, column=4).value = ar.get_attribute('href')
			except:
				pass
			sh.cell(row=i, column=1).value = i
			sh.cell(row=i, column=3).value = td[0].text
			i += 1
		
	wb.save(xli + '.xlsx')		
	print '(+) Saved ' + xli + '.xlsx'
	
