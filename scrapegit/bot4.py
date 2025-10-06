import selenium
from selenium import webdriver 
import openpyxl
import time
import random
import os
import docx

dl = os.listdir('/home/diablo/Desktop/scrpy/exxm/')
dr = webdriver.Chrome(r'/home/diablo/chromedriver')
for dx in dl:
	wb = openpyxl.load_workbook('/home/diablo/Desktop/scrpy/exxm/'+dx)
	print(dx)
	sh = wb.active
	c = sh.max_row
	print(c)
	i = 1
	
	for x in range(1, c+1):
		url = sh.cell(row=x,column=1).value
		print(x)
		try:
			dr.get(url)
		except:
			continue
		if not os.path.exists('/home/diablo/Desktop/scrpy/exm2/'+dx):
			os.mkdir('/home/diablo/Desktop/scrpy/exm2/'+dx)
		wxx = open('/home/diablo/Desktop/scrpy/exm2/'+dx+'/'+sh.cell(row=x,column=2).value+'.html')
		zx = dr.find_element_by_id('article_section_div')
		zc = zx.get_attribute('outerHTML')
		try:
			ex1 = zx.find_elements_by_class_name('bodyslot')
			for ex1x in ex1:
				zc = zc.replace(ex1x.get_attribute('outerHTML'), '')
		except:
			pass
		try:
			ex2 = zx.find_elements_by_tag_name('center')
			for ex2x in ex2:
				zc = zc.replace(ex2x.get_attribute('outerHTML'), '')
		except:
			pass
		
		wxx.write(zc)
		wxx.close()
		i += 1
		time.sleep(random.randint(10,15))
