import os
import openpyxl
import docx
from selenium import webdriver
pz = 0
prox = []

def rebro():
	print('$_$ Setting Proxy $_$')
	pth = r'/home/diablo/chromedriver'
	global pz
	global prox
	if len(prox) == 0:
		dri = webdriver.Chrome(pth)
		dri.get('https://free-proxy-list.net/')
		proxt = dri.find_elements_by_xpath('//*[@id="proxylisttable"]/tbody/tr')
		for x in proxt:
			row_data = x.find_elements_by_tag_name('td')
			proxy = row_data[0].text+":"+row_data[1].text
			prox.append(proxy)
		dri.quit()	
	px = prox[pz]
	options = webdriver.ChromeOptions()
	options.add_argument("--start-maximized")
	options.add_argument('--proxy-server=http://%s' % px)

	driver = webdriver.Chrome(pth,options=options)
	pz += 1
	print('$_$ Proxy Changed : '+prox[pz])
	return driver

br = rebro()

ld = os.listdir('/home/diablo/Desktop/dipl/')
for lx in ld:
	op = openpyxl.load_workbook('/home/diablo/Desktop/dipl/'+lx)
	oq = op.active
	ox = oq.max_row

	x = openpyxl.Workbook()
	xx = x.active
	xxx = 1
	
	for om in range(1,ox+1):

		y = open('/home/diablo/Desktop/diplo/'+str(om)+lx+'.html', 'w')
		br.get(oq.cell(row=om,column=2).value)
		
		pb = br.find_element_by_class_name('panel-body')
		h1 = pb.find_element_by_tag_name('h1')
		wl = pb.find_element_by_class_name('well')
		wlx = wl.find_elements_by_tag_name('tr')
		
		xx.cell(row=xxx,column=1).value = h1.text
		
		xxy = 2
		for wly in wlx:
			td = wly.find_elements_by_tag_name('td')
			xx.cell(row=xxx,column=xxy).value = td[-1].text
			xxy += 1
			x.save('/home/diablo/Desktop/diplo/'+lx+'_full.xlsx')
		pbr = pb.get_attribute('outerHTML')
		pbr = pbr.encode('ascii','ignore')
		y.write(pbr)
		y.close()