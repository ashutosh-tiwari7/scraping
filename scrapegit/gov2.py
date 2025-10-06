from selenium import webdriver
import docx
import openpyxl
import time

#start PhantomJS
#service_args = [ '--proxy=localhost:9150', '--proxy-type=socks5', ]
#driver = webdriver.PhantomJS(service_args=service_args)
driver = webdriver.Chrome(r'/usr/local/bin/chromedriver')
#alli = ['s', 't', 'u', 'v', 'w', 'x', 'y', 'z']
#doc = docx.Document()

xl = openpyxl.Workbook()
yl = openpyxl.Workbook()
zl = openpyxl.Workbook()
sh = xl.active
th = yl.active
uh = zl.active

i = 1
for ix in range(1,52):
	driver.get('http://educationportal.mp.gov.in/Public/Search/Search_DDO.aspx')
	
	print '(=)(=)(=)    Site fetched'

	sl = driver.find_element_by_id('ctl00_ctl00_RMCPH_ContentPlaceHolder1_ddlDistrict')
	bt = driver.find_element_by_id('ctl00_ctl00_RMCPH_ContentPlaceHolder1_btn_Search')
	
	op = sl.find_elements_by_tag_name('option')
	oq = iter(op)
	ot = next(oq)
	
	for om in range(1, i+1):	
		ot = next(oq)
	ot.click()
	dil = {}
	dil[i] = ot.text
	bt.click()

	time.sleep(1)

	tb = driver.find_element_by_id('ctl00_ctl00_RMCPH_ContentPlaceHolder1_GVPADistrict')
	
	
	tr = tb.find_elements_by_tag_name('tr')
	r1 = 1
	print '(+)(+)(+)   Printing rows'
	for trr in tr:
		try:
			td = trr.find_elements_by_tag_name('td')
		except:
			td = trr.find_elements_by_tag_name('th')
		
		r2 = 1
		for tx in td:	
			sh.cell(row = r1, column = r2).value = tx.text
			r2 += 1
		r1 += 1
		
	xl.save(dil[i]+'.xlsx')
	
	print '(+)(+)(+)   saved as' + dil[i]
	i += 1



