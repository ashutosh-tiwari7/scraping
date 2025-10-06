import selenium
import openpyxl
from selenium import webdriver
import time

driver = webdriver.Chrome(r'/home/diablo/chromedriver')
wb = openpyxl.Workbook()
sh = wb.active
i = 1
driver.maximize_window()
for t in range(0,3):
	driver.get('https://www.buddy4study.com/scholarships')
	bp = driver.find_element_by_class_name('buttonpanel')
	sp = bp.find_elements_by_tag_name('span')

	sp[t].click()
	time.sleep(2)
	
	for x in range(1, 5000):
		
		nm = driver.find_element_by_class_name('flex-container-list')
		nn = nm.find_elements_by_class_name('ellipsis')
		try:
			snn = driver.find_element_by_class_name('fix-banner')
			snm = snn.find_element_by_class_name('cross')
			snm.click()
		except:
			pass
		try:
			driver.execute_script("var x = document.getElementsByClassName('stickyHaderWrapper'); x[0].style.display = 'none';")
		except:
			pass
		for nx in nn:
			sh.cell(row = i, column = 1).value = nx.text
			sh.cell(row = i, column = 2).value = nx.get_attribute('href')
			i += 1

		wb.save('bfs.xlsx')
		
		btn = driver.find_element_by_class_name('rc-pagination-next')
		atb = btn.get_attribute('disabled')
		
		if atb == 'true':
			break
		else:
			
			btn.click()
			








