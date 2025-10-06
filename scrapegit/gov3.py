from selenium import webdriver
import docx
import openpyxl
import time
import urllib

#start PhantomJS
#service_args = [ '--proxy=localhost:9150', '--proxy-type=socks5', ]
#driver = webdriver.PhantomJS(service_args=service_args)
driver = webdriver.Chrome(r'/usr/local/bin/chromedriver')
#alli = ['s', 't', 'u', 'v', 'w', 'x', 'y', 'z']
#doc = docx.Document()

xl = openpyxl.Workbook()
sh = xl.active
zl = openpyxl.Workbook()
zh = zl.active

k = 9
u = 19

driver.get('http://educationportal.mp.gov.in/ESB/Public/BloodGroup.aspx')

sl1 = driver.find_element_by_id('DDL_Blood_Group_ID')
sl2 = driver.find_element_by_id('DDLDistrict')

op1 = sl1.find_elements_by_tag_name('option')
op2 = sl2.find_elements_by_tag_name('option')

oq1 = iter(op1)
oq2 = iter(op2)

ot1 = next(oq1)
ot2 = next(oq2)

for xz in range(1, k+1):
	ot1 = next(oq1)
	
for yz in range(1, u+1):
	ot2 = next(oq2)	

ot1.click()
ot2.click()

btn = driver.find_element_by_id('BtnView')
btn.click()

for fx in range(1,50000):
	
	i = 2
	t = 1
	for fz in range(1,50000):
		tbl = driver.find_element_by_id('GrdEmployee')
		
		trs = tbl.find_elements_by_tag_name('tr')
		tr = tbl.find_element_by_tag_name('tr')
		print '(+) Rows found : ' + str(len(trs))
		
		for trx in trs[3:-2]:
			tdd = trx.find_elements_by_tag_name('td')
			print '(+) Data Cells : ' + str(len(tdd))
			
			#img = tdd[0].find_element_by_tag_name('img')
			c1 = tdd[1].text
			c2 = c1.split('[')
			c3 = c2[1].split(']')
			d1 = tdd[3].text
			d2 = d1.split('District')
			dist = d2[1]

			sh.cell(row = t, column = 1).value = c2[0]
			sh.cell(row = t, column = 2).value = c3[0]
			sh.cell(row = t, column = 3).value = c3[1]
			sh.cell(row = t, column = 4).value = tdd[2].text
			sh.cell(row = t, column = 5).value = d2[0]
			sh.cell(row = t, column = 6).value = d2[1]
			sh.cell(row = t, column = 7).value = tdd[4].text
			sh.cell(row = t, column = 8).value = tdd[5].text
			sh.cell(row = t, column = 9).value = tdd[6].text

			zh.cell(row = t, column = 1).value = i
			zh.cell(row = t, column = 2).value = c2[0]
			zh.cell(row = t, column = 3).value = d2[1]
			zh.cell(row = t, column = 4).value = tdd[4].text
			zh.cell(row = t, column = 5).value = tdd[5].text
			zh.cell(row = t, column = 6).value = tdd[6].text
			t += 1
		
		xl.save('group' + str(k) + str(u) + 'all.xlsx')
		zl.save('group' + str(k) + str(u) + 'sub.xlsx')
		
		print '(+)(++) Saved As : page: ' + str(i) + '; group: ' + str(k) + '; city: ' + str(u) + ' all.xlsx'
		print '(+)(++) Saved As : page: ' + str(i) + '; group: ' + str(k) + '; city: ' + str(u) + ' sub.xlsx'
		try:
			trd = tr.find_element_by_link_text(str(i))
		except:
			trm = tr.find_elements_by_link_text('...')
			trd = trm[-1]
		i += 1
		trd.click()
		
		#qazwsx&@.@25edc
		
		
#guna gwalior 
#2 - 1791
#3 - 1955
# 5 - 1737
# 4 - 1753

