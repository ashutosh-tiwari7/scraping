from selenium import webdriver
import docx
import openpyxl
import time

driver = webdriver.Chrome(r'/usr/local/bin/chromedriver')
driver.maximize_window()
sh = openpyxl.Workbook()
wb = sh.active
ll = ['https://yellowpages.webindia123.com/search_pro.aspx?city=Gwalior&q=coaching&street=', 
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Indore&q=coaching&street=Street',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Ahmedabad&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Allahabad&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Ajmer&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Amritsar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Bangalore&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Bhubaneswar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Bareilly&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Chandigarh&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Coimbatore&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Chennai&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Dehradun&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Delhi&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Faridabad&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Gandhinagar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Jaipur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Jalandhar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Jamshedpur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Jamnagar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Kanpur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Kolkata&q=coaching&street='
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Kochi&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Kanyakumari&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Lucknow&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Ludhiana&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Madurai&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Meerut&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Mumbai&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Mangalore&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Nagpur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Nasik&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Noida&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Goa&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Gurgaon&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Guwahati&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Hyderabad&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Howrah&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Patiala&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Patna&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Pune&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Raipur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Ranchi&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Shimla&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Srinagar&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Thane&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Thiruvananthapuram&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Udaipur&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Vadodara&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Visakhapatnam&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Varanasi&q=coaching&street=',
	  'https://yellowpages.webindia123.com/search_pro.aspx?city=Vellore&q=coaching&street=']
z = 1
for lx in ll:
	driver.get(lx)
	i = 1
	j = 1

	for ty in range(1, 500000):

		tb = driver.find_elements_by_id('bg_basic')

		if driver.find_elements_by_id('spon'):
			sp = driver.find_elements_by_id('spon')
			for sx in sp:
				wb.cell(row = i, column = 1).value = sx.text
				wb.cell(row = i, column = 2).value = sx.get_attribute('href')
				i += 1

		for tx in tb:
			a = tx.find_element_by_id('basic')
			wb.cell(row = i, column = 1).value = a.text
			wb.cell(row = i, column = 2).value = a.get_attribute('href')
			i += 1	

		sh.save(str(z) + '.xlsx')
		j += 1
		try:
			b = driver.find_element_by_link_text(str(j))
			b.click()
		except:
			break
	z += 1












