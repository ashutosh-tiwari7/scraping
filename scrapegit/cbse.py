from selenium import webdriver
import docx
import openpyxl
import time
import urllib
from selenium.webdriver.common.keys import Keys

#start PhantomJS
#service_args = [ '--proxy=localhost:9150', '--proxy-type=socks5', ]
#driver = webdriver.PhantomJS(service_args=service_args)
driver = webdriver.Chrome(r'/usr/local/bin/chromedriver')
#alli = ['s', 't', 'u', 'v', 'w', 'x', 'y', 'z']
#doc = docx.Document()

xl = openpyxl.Workbook()
sh = xl.active
driver.maximize_window()
k = 32

driver.get('http://cbseaff.nic.in/cbse_aff/schdir_Report/userview.aspx')

ot1 = driver.find_element_by_id('optlist_2')
ot1.click()

#time.sleep(3)

ot2 = driver.find_element_by_id('ddlitem')

op1 = ot2.find_elements_by_tag_name('option')

op2 = iter(op1)
op = next(op2)

for z in range(1, k+1):
	op = next(op2)

stt = op.text	
op.click()
time.sleep(1)

print '(+)(+) Printing State : ' + stt

sb = driver.find_element_by_id('search')
sb.click()
i = 1

for sx in range(1, 50000):
	
	tbb = driver.find_element_by_id('T1')
	hyr = tbb.find_elements_by_tag_name('a')
	tot = driver.find_element_by_id('tot')
	
	print '(+)(+) Rows Found : ' + str(len(hyr))
	print '(+)(+) Page : ' + str(sx)

	for trx in hyr:
		td = trx.get_attribute('href')
		td2 = trx.text

		sh.cell(row = i, column = 1).value = i
		sh.cell(row = i, column = 2).value = td
		sh.cell(row = i, column = 3).value = td2
		i += 1

	xl.save(stt + '.xlsx')
	print '(+)(+)(+)(+)(+) Saved'
	
	driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
	
	nx = driver.find_element_by_id('Button1')
	nx.click()

	